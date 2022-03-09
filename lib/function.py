from openpyxl import load_workbook
import os
import pandas as pd
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
    writer.book = load_workbook(filename)
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    if truncate_sheet and sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
        writer.book.create_sheet(sheet_name, idx)

    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
